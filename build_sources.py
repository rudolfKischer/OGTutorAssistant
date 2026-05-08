"""
Download and parse external data sources into compact JSON files.

This script reconstructs three data files from their upstream sources:
  1. g2p_aligned.json   — grapheme-phoneme alignments (from aligned-cmudict)
  2. wiktionary_parsed.json — definitions + POS (from kaikki.org Wiktionary extract)
  3. morpholex_parsed.json  — morpheme segmentation (from MorphoLex-en)

Run:  python3 build_sources.py [g2p|wiktionary|morpholex|all]
"""

import gzip
import json
import os
import re
import sys
import urllib.request

from config import DATA_DIR, G2P_ALIGNED_PATH, WIKTIONARY_PATH, MORPHOLEX_JSON_PATH

SOURCES = {
    'g2p': {
        'url': 'https://raw.githubusercontent.com/ckw017/aligned-cmudict/master/g2p.json',
        'description': 'Phonetisaurus grapheme-phoneme alignments for CMU dict',
        'repo': 'https://github.com/ckw017/aligned-cmudict',
    },
    'wiktionary': {
        'url': 'https://kaikki.org/dictionary/raw-wiktextract-data.jsonl.gz',
        'description': 'English Wiktionary extract (definitions, POS, examples)',
        'site': 'https://kaikki.org/dictionary/rawdata.html',
    },
    'morpholex': {
        'url': 'https://raw.githubusercontent.com/hugomailhot/MorphoLex-en/master/MorphoLEX_en.xlsx',
        'description': 'MorphoLex morphological segmentation database',
        'repo': 'https://github.com/hugomailhot/MorphoLex-en',
    },
}


def _download(url, dest, label):
    if os.path.exists(dest):
        print(f"  {label} already downloaded: {dest}")
        return
    print(f"  Downloading {label}...")
    urllib.request.urlretrieve(url, dest)
    size_mb = os.path.getsize(dest) / 1e6
    print(f"  Downloaded {size_mb:.1f} MB -> {dest}")


# --- G2P aligned ---

def build_g2p():
    """Download and compact g2p.json -> g2p_aligned.json."""
    print("\n=== G2P Alignments ===")
    print(f"  Source: {SOURCES['g2p']['url']}")

    raw_path = os.path.join(DATA_DIR, '_g2p_raw.json')
    _download(SOURCES['g2p']['url'], raw_path, 'g2p.json')

    print("  Parsing...")
    with open(raw_path, 'r') as f:
        raw = json.load(f)

    compact = {}
    for word, entry in raw.items():
        pairs = []
        for g, p in zip(entry['graphemes'], entry['phonemes']):
            clean_g = g.replace('|', '')
            clean_p = p.rstrip('012') if p != '_' else ''
            pairs.append([clean_g, clean_p])
        compact[word] = pairs

    with open(G2P_ALIGNED_PATH, 'w') as f:
        json.dump(compact, f, ensure_ascii=False, separators=(',', ':'))

    os.remove(raw_path)
    size_mb = os.path.getsize(G2P_ALIGNED_PATH) / 1e6
    print(f"  Wrote {len(compact)} words -> {G2P_ALIGNED_PATH} ({size_mb:.1f} MB)")


# --- Wiktionary ---

POS_MAP = {
    'noun': 'noun', 'verb': 'verb', 'adj': 'adj', 'adv': 'adv',
    'pron': 'pron', 'det': 'det', 'prep': 'adp', 'adp': 'adp',
    'conj': 'conj', 'num': 'num', 'particle': 'prt', 'prt': 'prt',
    'intj': 'intj', 'contraction': 'contraction',
    'article': 'det', 'prefix': 'noun', 'suffix': 'noun',
    'name': 'noun', 'phrase': None, 'proverb': None,
    'character': 'noun', 'symbol': 'noun', 'abbrev': None,
}


def _parse_wikt_entry(entry):
    """Extract (pos, definitions) from a kaikki.org entry."""
    raw_pos = entry.get('pos', '').lower()
    pos = POS_MAP.get(raw_pos, raw_pos if len(raw_pos) <= 5 else None)
    if not pos:
        return []

    results = []
    for sense in entry.get('senses', []):
        glosses = sense.get('glosses', [])
        if not glosses:
            continue
        definition = glosses[-1]
        if definition.startswith('Alternative') or definition.startswith('Obsolete'):
            continue

        example = None
        for ex in sense.get('examples', []):
            text = ex.get('text', '')
            if text and len(text) < 300:
                example = text
                break

        results.append([pos, definition, example or ''])
    return results


def build_wiktionary():
    """Download kaikki.org Wiktionary dump -> wiktionary_parsed.json."""
    print("\n=== Wiktionary ===")
    print(f"  Source: {SOURCES['wiktionary']['url']}")

    gz_path = os.path.join(DATA_DIR, '_wiktionary_raw.jsonl.gz')
    _download(SOURCES['wiktionary']['url'], gz_path, 'wiktionary JSONL')

    print("  Parsing English entries (this takes a few minutes)...")
    wikt = {}
    total_lines = 0
    with gzip.open(gz_path, 'rt', encoding='utf-8') as f:
        for line in f:
            total_lines += 1
            try:
                entry = json.loads(line)
            except json.JSONDecodeError:
                continue

            if entry.get('lang_code') != 'en':
                continue

            word = entry.get('word', '').lower().strip()
            if not word or not all(c.isalpha() or c in "-'" for c in word):
                continue

            defs = _parse_wikt_entry(entry)
            if defs:
                wikt.setdefault(word, []).extend(defs)

    # Deduplicate definitions per word
    for word in wikt:
        seen = set()
        unique = []
        for d in wikt[word]:
            key = d[1]  # definition text
            if key not in seen:
                seen.add(key)
                unique.append(d)
        wikt[word] = unique

    with open(WIKTIONARY_PATH, 'w') as f:
        json.dump(wikt, f, ensure_ascii=False, separators=(',', ':'))

    os.remove(gz_path)
    total_defs = sum(len(v) for v in wikt.values())
    size_mb = os.path.getsize(WIKTIONARY_PATH) / 1e6
    print(f"  Processed {total_lines:,} lines")
    print(f"  Wrote {len(wikt):,} words, {total_defs:,} definitions -> {WIKTIONARY_PATH} ({size_mb:.1f} MB)")


# --- MorphoLex ---

def _parse_morpholex_sheet(ws):
    """Extract {word: segmentation} from an openpyxl worksheet."""
    results = {}
    headers = None
    for row in ws.iter_rows(values_only=True):
        vals = list(row)
        if headers is None:
            if 'Word' in vals:
                headers = vals
            continue
        word = vals[headers.index('Word')]
        segm = vals[headers.index('MorphoLexSegm')]
        if isinstance(word, str) and isinstance(segm, str) and word and segm:
            results[word.lower()] = segm
    return results


def build_morpholex():
    """Download MorphoLex Excel -> morpholex_parsed.json."""
    print("\n=== MorphoLex ===")
    print(f"  Source: {SOURCES['morpholex']['url']}")

    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl required. Install with: pip install openpyxl")
        return

    xlsx_path = os.path.join(DATA_DIR, '_MorphoLEX_en.xlsx')
    _download(SOURCES['morpholex']['url'], xlsx_path, 'MorphoLEX_en.xlsx')

    print("  Parsing Excel sheets...")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    skip_sheets = {'Presentation', 'Prefixes', 'Suffixes', 'Roots'}

    raw = {}
    for sname in wb.sheetnames:
        if sname not in skip_sheets:
            raw.update(_parse_morpholex_sheet(wb[sname]))
    wb.close()

    # Parse segmentation strings into [prefixes, roots, suffixes]
    compact = {}
    for word, segm in raw.items():
        compact[word] = [
            re.findall(r'<([a-zA-Z]+)<', segm),
            re.findall(r'\(([a-zA-Z]+)\)', segm),
            re.findall(r'>([a-zA-Z]+)>', segm),
        ]

    with open(MORPHOLEX_JSON_PATH, 'w') as f:
        json.dump(compact, f, ensure_ascii=False, separators=(',', ':'))

    os.remove(xlsx_path)
    size_mb = os.path.getsize(MORPHOLEX_JSON_PATH) / 1e6
    print(f"  Wrote {len(compact)} words -> {MORPHOLEX_JSON_PATH} ({size_mb:.1f} MB)")


# --- Main ---

BUILDERS = {
    'g2p': build_g2p,
    'wiktionary': build_wiktionary,
    'morpholex': build_morpholex,
}


def main():
    os.makedirs(DATA_DIR, exist_ok=True)

    targets = sys.argv[1:] or ['all']
    if 'all' in targets:
        targets = list(BUILDERS.keys())

    for target in targets:
        if target not in BUILDERS:
            print(f"Unknown target: {target}")
            print(f"Usage: python3 build_sources.py [{' | '.join(BUILDERS)} | all]")
            sys.exit(1)
        BUILDERS[target]()

    print("\nDone. Next steps:")
    print("  python3 build_word_data.py   # build words.json")
    print("  python3 build_word_db.py     # build words.db")


if __name__ == '__main__':
    main()
